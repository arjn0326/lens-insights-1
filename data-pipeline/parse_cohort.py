"""
parse_cohort.py
---------------
Extracts parish-level cohort graduation and credential rates.

Inputs (place in data-pipeline/raw/):
  2024-state-school-system-and-school-cohort-grad-rates-by-subgroups.xlsx
  2025-state-school-system-school-cohort-credential-rates-by-subgroups.xlsx

Outputs (written to public/data/):
  cohort_by_parish.json
  cohort_by_parish.csv

Usage:
  cd data-pipeline
  python parse_cohort.py
"""

import json
import csv
import re
from pathlib import Path

try:
    import openpyxl
except ImportError:
    raise SystemExit("openpyxl not found. Run: pip install openpyxl")

PIPELINE_DIR = Path(__file__).parent
RAW_DIR      = PIPELINE_DIR / "raw"
OUT_DIR      = PIPELINE_DIR.parent / "public" / "data"

GRAD_FILE = RAW_DIR / "2024-state-school-system-and-school-cohort-grad-rates-by-subgroups.xlsx"
CRED_FILE = RAW_DIR / "2025-state-school-system-school-cohort-credential-rates-by-subgroups.xlsx"

def clean_val(val):
    if val is None:
        return None
    s = str(val).strip()
    if s in ("", "None"):
        return None
    if s in ("NR", "~", ">95", "<5"):
        return s
    try:
        return round(float(s), 1)
    except ValueError:
        return s

def is_parish(code):
    if code == "LA":
        return True
    if re.match(r"^\d{3}$", str(code)):
        n = int(code)
        return 1 <= n <= 64
    return False

def equity_gap(white, black):
    if isinstance(white, float) and isinstance(black, float):
        return round(white - black, 1)
    return None

def parish_slug(name):
    slug = name.lower()
    slug = slug.replace(" parish", "").replace("louisiana statewide", "louisiana").strip()
    slug = re.sub(r"[^a-z0-9]+", "-", slug).strip("-")
    return slug

# ── File 1: Graduation rates ───────────────────────────────────────────────
# Headers at row 4, data from row 5
# Col 0=code, 1=name, 2=overall, 3=am.indian, 4=asian,
# 5=black, 6=hispanic, 7=white, 8=hawaiian, 9=multirace,
# 10=econ_dis, 11=swd, 12=el, 13=homeless, 14=foster, 15=military

print("Reading graduation rates file...")
wb1 = openpyxl.load_workbook(str(GRAD_FILE), read_only=True)
ws1 = wb1.active
grad_data = {}

for row in ws1.iter_rows(min_row=5, values_only=True):
    code = str(row[0]).strip() if row[0] else ""
    if not is_parish(code):
        continue
    name = str(row[1]).strip() if row[1] else ""
    grad_data[code] = {
        "parish_code":                  code,
        "parish_name":                  name,
        "parish_slug":                  parish_slug(name),
        "grad_overall":                 clean_val(row[2]),
        "grad_black":                   clean_val(row[5]),
        "grad_hispanic":                clean_val(row[6]),
        "grad_white":                   clean_val(row[7]),
        "grad_econ_disadvantaged":      clean_val(row[10]),
        "grad_students_w_disabilities": clean_val(row[11]),
        "grad_english_learner":         clean_val(row[12]),
    }

wb1.close()
print(f"  Loaded: {len(grad_data)} rows")

# ── File 2: Credential rates ───────────────────────────────────────────────
# Headers at row 4, data from row 5
# Col 0=code, 1=name, 2=subgroup,
# 3=advanced, 4=basic, 5=combined, 6=no_credentials
# Only keep subgroup == "Total Population"

print("Reading credential rates file...")
wb2 = openpyxl.load_workbook(str(CRED_FILE), read_only=True)
ws2 = wb2.active
cred_data = {}

for row in ws2.iter_rows(min_row=5, values_only=True):
    code = str(row[0]).strip() if row[0] else ""
    if not is_parish(code):
        continue
    subgroup = str(row[2]).strip() if row[2] else ""
    if subgroup != "Total Population":
        continue
    cred_data[code] = {
        "cred_advanced":       clean_val(row[3]),
        "cred_basic":          clean_val(row[4]),
        "cred_combined":       clean_val(row[5]),
        "cred_no_credentials": clean_val(row[6]),
    }

wb2.close()
print(f"  Loaded: {len(cred_data)} rows")

# ── Merge ──────────────────────────────────────────────────────────────────
all_codes = sorted(set(list(grad_data.keys()) + list(cred_data.keys())))
parishes  = []

for code in all_codes:
    g = grad_data.get(code, {})
    c = cred_data.get(code, {})
    parishes.append({
        "parish_code":                  code,
        "parish_name":                  g.get("parish_name",  c.get("parish_name",  "")),
        "parish_slug":                  g.get("parish_slug",  c.get("parish_slug",  "")),
        "grad_overall":                 g.get("grad_overall"),
        "grad_black":                   g.get("grad_black"),
        "grad_hispanic":                g.get("grad_hispanic"),
        "grad_white":                   g.get("grad_white"),
        "grad_econ_disadvantaged":      g.get("grad_econ_disadvantaged"),
        "grad_students_w_disabilities": g.get("grad_students_w_disabilities"),
        "grad_english_learner":         g.get("grad_english_learner"),
        "equity_gap_white_minus_black": equity_gap(
            g.get("grad_white"), g.get("grad_black")
        ),
        "cred_advanced":                c.get("cred_advanced"),
        "cred_basic":                   c.get("cred_basic"),
        "cred_combined":                c.get("cred_combined"),
        "cred_no_credentials":          c.get("cred_no_credentials"),
    })

state_row   = next((p for p in parishes if p["parish_code"] == "LA"), None)
parish_rows = [p for p in parishes if p["parish_code"] != "LA"]
print(f"  Merged: {len(parish_rows)} parishes + 1 state row")

# ── Write JSON ─────────────────────────────────────────────────────────────
out_json = OUT_DIR / "cohort_by_parish.json"
with open(out_json, "w") as f:
    json.dump({
        "meta": {
            "source_grad":      "LDOE 2023-2024 Cohort Graduation Rates by Subgroup",
            "source_cred":      "LDOE 2024-2025 Cohort Credential Rates by Subgroup",
            "notes": {
                "NR":   "Not reported — subgroup too small or data unavailable",
                "~":    "Statistically unreliable — fewer than 10 students",
                ">95":  "Greater than 95% — exact value suppressed for privacy",
                "<5":   "Less than 5% — exact value suppressed for privacy",
            },
            "equity_gap_note":  "White grad rate minus Black grad rate. Positive = White students graduate at higher rate.",
            "cred_advanced_note": "Advanced = 150+ credits including rigorous coursework, CTE, dual enrollment, or AP. Workforce/college ready.",
            "cred_no_cred_note":  "Diploma with no credentials = minimum graduation only. Not workforce or college ready.",
        },
        "state": state_row,
        "parishes": parish_rows,
    }, f, indent=2)
print(f"JSON → {out_json}")

# ── Write CSV ──────────────────────────────────────────────────────────────
out_csv = OUT_DIR / "cohort_by_parish.csv"
fields = [
    "parish_code", "parish_name", "parish_slug",
    "grad_overall", "grad_black", "grad_hispanic", "grad_white",
    "grad_econ_disadvantaged", "grad_students_w_disabilities",
    "grad_english_learner", "equity_gap_white_minus_black",
    "cred_advanced", "cred_basic", "cred_combined", "cred_no_credentials",
]

all_rows = ([state_row] if state_row else []) + parish_rows
with open(out_csv, "w", newline="") as f:
    w = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
    w.writeheader()
    w.writerows(all_rows)
print(f"CSV  → {out_csv}")
print("Done.")
