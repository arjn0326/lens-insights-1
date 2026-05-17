import openpyxl

files = [
    "raw/2024-state-school-system-and-school-cohort-grad-rates-by-subgroups.xlsx",
    "raw/2025-state-school-system-school-cohort-credential-rates-by-subgroups.xlsx",
]

for filepath in files:
    print(f"\n{'='*60}")
    print(f"FILE: {filepath}")
    wb = openpyxl.load_workbook(filepath, read_only=True)
    print(f"Sheets: {wb.sheetnames}")
    ws = wb.active
    print("First 8 rows:")
    for i, row in enumerate(ws.iter_rows(max_row=8, values_only=True)):
        print(f"  Row {i+1}: {row[:10]}")
    wb.close()